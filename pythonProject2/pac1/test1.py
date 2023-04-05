import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pytest
from pac1.pageobject1 import registration1

@pytest.fixture()
def setup():
    print("this runs before every method")


def testdef1(setup):
    print("Hello world")


def test_registeropencart1(setup):
    excelpath="C:\\Users\\Administrator\\Documents\\python project\\registrationdataimport.xlsx"
    workbook=openpyxl.load_workbook(excelpath)
    sheet=workbook.active
    rows=sheet.max_row
    cols=sheet.max_column
    print(rows)
    for i in range(2,rows+1):
        fname=sheet.cell(row=i,column=1).value
        lname=sheet.cell(row=i, column=2).value
        mail=sheet.cell(row=i, column=3).value
        passw=sheet.cell(row=i, column=4).value
        #loginopencart(fname,lname,mail,passw)
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get("https://demo.opencart.com/")
        driver.maximize_window()
        rp = registration1(driver)
        rp.clikmyaccountlink()
        rp.clickonregisterlink()
        time.sleep(5)
        rp.inputfirstname(fname)
        rp.inputlastname(lname)
        rp.enteremail(mail)
        rp.enterpassword(passw)
        rp.checkbox()
        time.sleep(5)
        rp.clickonsubmit()
        time.sleep(5)
