import time
import openpyxl as openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
#from webdriver_manager.core import driver

def loginopencart(firstname,lastname,email,password):
    driver=webdriver.Chrome(ChromeDriverManager().install())
    driver.get("https://demo.opencart.com/")
    driver.find_element(By.LINK_TEXT,"My Account").click()
    driver.find_element(By.LINK_TEXT,"Register").click()
    time.sleep(5)
    print(driver.title)
    if driver.title=="Your Store":
        print("title is matched")
    else:
        print("title is not matched")
    driver.find_element(By.ID,"input-firstname").send_keys(firstname)
    driver.find_element(By.ID,"input-lastname").send_keys(lastname)
    driver.find_element(By.ID,"input-email").send_keys(email)
    driver.find_element(By.ID,"input-password").send_keys(password)
    driver.find_element(By.NAME,"agree").click()
    driver.find_element(By.XPATH,"//button[@type='submit']").click()
    time.sleep(5)
excelpath="C:\\Users\\Administrator\\Documents\\python project\\registrationdataimport.xlsx"
workbook=openpyxl.load_workbook(excelpath)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
print(rows)
for i in range(2,rows+1):
    fname=sheet.cell(row=i,column=1).value
    lname=sheet.cell(row=i, column=2).value
    mail=sheet.cell(row=i, column=3).value
    passw=sheet.cell(row=i, column=4).value
    loginopencart(fname,lname,mail,passw)